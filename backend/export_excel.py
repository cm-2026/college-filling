#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
高考志愿推荐导出Excel脚本
"""

import sys
import json
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

def create_excel(data, output_path):
    """创建Excel文件"""
    wb = Workbook()
    ws = wb.active
    ws.title = '志愿推荐'
    
    # 获取数据
    region = data.get('region', '')
    score = data.get('score', '')
    rank = data.get('rank', '')
    subject = data.get('subject', '')
    schools = data.get('schools', [])
    
    # 计算统计
    school_count = len(schools)
    major_count = sum(len(s.get('majors', [])) for s in schools)
    
    # 样式定义
    bold_font = Font(bold=True)
    center_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 第一行：生源地、分数、位次
    ws['A1'] = '生源地'
    ws['B1'] = region
    ws['C1'] = '分数'
    ws['D1'] = score
    ws['E1'] = '位次'
    ws['F1'] = rank
    
    # 第二行：选科、院校数量、专业数量
    ws['A2'] = '选科'
    ws['B2'] = subject
    ws['C2'] = '院校数量'
    ws['D2'] = school_count
    ws['E2'] = '专业数量'
    ws['F2'] = major_count
    
    # 应用样式到前两行
    for row in range(1, 3):
        for col in range(1, 7):
            cell = ws.cell(row=row, column=col)
            cell.alignment = center_align
            cell.border = thin_border
            if col in [1, 3, 5]:  # 标签列加粗
                cell.font = bold_font
    
    # 设置列宽
    ws.column_dimensions['A'].width = 35.625
    ws.column_dimensions['B'].width = 13
    ws.column_dimensions['C'].width = 13
    ws.column_dimensions['D'].width = 13
    ws.column_dimensions['E'].width = 13
    ws.column_dimensions['F'].width = 13
    
    # 当前行
    current_row = 3
    
    # 遍历每个学校
    for school in schools:
        school_name = school.get('name', '')
        majors = school.get('majors', [])
        
        if not majors:
            continue
        
        # 写表头
        ws.cell(row=current_row, column=1, value='院校名称')
        ws.cell(row=current_row, column=2, value='专业名称')
        ws.cell(row=current_row, column=3, value='专业类')
        ws.cell(row=current_row, column=4, value='最低分')
        ws.cell(row=current_row, column=5, value='最低位次')
        
        # 表头样式
        for col in range(1, 6):
            cell = ws.cell(row=current_row, column=col)
            cell.font = bold_font
            cell.alignment = center_align
            cell.border = thin_border
        
        current_row += 1
        
        # 写专业数据
        start_row = current_row
        for i, major in enumerate(majors):
            major_name = major.get('major', '')
            major_category = major.get('major_category', '')
            min_score = major.get('min_score', '') or major.get('group_min_score_1', '')
            min_rank = major.get('min_rank', '') or major.get('group_min_rank_1', '')
            
            # 院校名称只在第一行写
            if i == 0:
                ws.cell(row=current_row, column=1, value=school_name)
            
            ws.cell(row=current_row, column=2, value=major_name)
            ws.cell(row=current_row, column=3, value=major_category)
            ws.cell(row=current_row, column=4, value=min_score)
            ws.cell(row=current_row, column=5, value=min_rank)
            
            # 数据样式
            for col in range(1, 6):
                cell = ws.cell(row=current_row, column=col)
                cell.alignment = center_align
                cell.border = thin_border
            
            current_row += 1
        
        # 合并院校名称单元格
        if len(majors) > 1:
            ws.merge_cells(start_row=start_row, start_column=1, end_row=current_row-1, end_column=1)
    
    # 确保目录存在
    output_dir = os.path.dirname(output_path)
    if output_dir and not os.path.exists(output_dir):
        os.makedirs(output_dir)
    
    # 保存文件
    wb.save(output_path)
    wb.close()  # 确保关闭工作簿
    
    # 验证文件
    if os.path.exists(output_path):
        file_size = os.path.getsize(output_path)
        print(f'[OK] Excel file generated: {output_path}, size: {file_size} bytes')
        return True
    else:
        print(f'[ERROR] Excel file generation failed: {output_path}')
        return False

def main():
    if len(sys.argv) < 3:
        print('用法: python export_excel.py <input_json> <output_xlsx>')
        sys.exit(1)
    
    input_path = sys.argv[1]
    output_path = sys.argv[2]
    
    print(f'[INFO] Reading data: {input_path}')
    
    # 读取JSON数据
    with open(input_path, 'r', encoding='utf-8') as f:
        data = json.load(f)
    
    print(f'[INFO] Data content: {len(data.get("schools", []))} schools')
    
    # 创建Excel
    success = create_excel(data, output_path)
    
    if success:
        sys.exit(0)
    else:
        sys.exit(1)

if __name__ == '__main__':
    main()
